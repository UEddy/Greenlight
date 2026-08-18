"""Parse data/raw/ into curated JSON under data/processed/.

Usage:
    python scripts/build-dataset.py

Run scripts/fetch-sources.py first. This script reads data/manifest.json for
provenance, so every output record can name the exact file it came from.

Design rules taken from the spec:
  - Every record carries source_url, source_year, axis, methodology,
    numerator and denominator. Where a source does not publish a numerator
    and denominator, they are null and the methodology says why. They are
    never reconstructed.
  - The two sources measure different things and are written to separate
    files. Nothing here blends them into one number.
  - Coverage gaps fail the build. If one of the 12 nationalities is missing,
    this script raises rather than quietly writing 11 records.
"""

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from financial_requirements import (  # noqa: E402
    AMOUNT_STATUSES,
    ANNEX_VERSION,
    BASES,
    RECORDS as FINANCIAL_RECORDS,
    RETRIEVED_AT as FINANCIAL_RETRIEVED_AT,
)
from sources import (  # noqa: E402
    COUNTRIES,
    FINANCIAL_THRESHOLD_CAVEAT,
    NATIONALITIES,
    SCHENGEN_CYPRUS_METHODOLOGY,
    SCHENGEN_FINANCIAL_METHODOLOGY,
    UK_FINANCIAL_METHODOLOGY,
    US_FINANCIAL_METHODOLOGY,
)

import openpyxl  # noqa: E402
from pypdf import PdfReader  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent
MANIFEST_PATH = REPO_ROOT / "data" / "manifest.json"
PROCESSED_DIR = REPO_ROOT / "data" / "processed"

UK_SOURCE_ID = "uk_home_office_vis_d02"
US_SOURCE_ID = "us_state_b_visa_adjusted_refusal_rates"
SCHENGEN_SOURCE_ID = "eu_commission_schengen_consulates"
FINANCIAL_SOURCE_ID = "eu_commission_reference_amounts_annex_25"

# The 29 states fully applying the Schengen acquis in 2026. Cyprus is not one
# of them and is carried separately, the same way the consulate dataset does
# it. If a state is missing from the curated table, the build fails rather
# than shipping a dataset with a silent hole in it.
SCHENGEN_STATES = [
    "Austria", "Belgium", "Bulgaria", "Croatia", "Czech Republic", "Denmark",
    "Estonia", "Finland", "France", "Germany", "Greece", "Hungary", "Iceland",
    "Italy", "Latvia", "Liechtenstein", "Lithuania", "Luxembourg", "Malta",
    "Netherlands", "Norway", "Poland", "Portugal", "Romania", "Slovakia",
    "Slovenia", "Spain", "Sweden", "Switzerland",
]

UK_TARGET_YEAR = "2025"
UK_VISA_GROUP = "Visitor"

# Population rate warning attached to every record. The spec calls this the
# single most important honesty rule, so it travels with the data rather than
# living only in the UI layer, where it could be dropped.
BASE_RATE_CAVEAT = (
    "This is a population base rate for a nationality group, not a personal "
    "probability. Do not render it as this applicant's odds."
)


def load_manifest():
    if not MANIFEST_PATH.exists():
        raise SystemExit(
            "data/manifest.json not found. Run python scripts/fetch-sources.py first."
        )
    manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    return {source["id"]: source for source in manifest["sources"]}


def raw_path(source):
    path = REPO_ROOT / source["raw_path"]
    if not path.exists():
        raise SystemExit(
            f"{source['raw_path']} not found. Run python scripts/fetch-sources.py first."
        )
    return path


def assert_full_coverage(found, dataset_name):
    """Fail the build if any of the 12 nationalities is missing."""
    missing = [n["name"] for n in NATIONALITIES if n["name"] not in found]
    if missing:
        raise SystemExit(
            f"{dataset_name}: no data found for {', '.join(missing)}. "
            f"The source label may have changed. Check the label columns in "
            f"scripts/sources.py against the raw file before editing this script."
        )


def build_uk(source):
    """UK Home Office Vis_D02, Visitor group, calendar year 2025."""
    wanted = {n["uk_label"]: n for n in NATIONALITIES}
    workbook = openpyxl.load_workbook(raw_path(source), read_only=True, data_only=True)
    sheet = workbook["Data_Vis_D02"]

    # Columns on Data_Vis_D02, header on row 4:
    # 0 Year, 1 Quarter, 2 Nationality, 3 Region, 4 Visa type group,
    # 5 Visa type, 6 Visa type subgroup, 7 Applicant type, 8 Case outcome,
    # 9 Decisions.
    # Within the Visitor group, applicant type is always All and visa type is
    # flat, so there is exactly one row per nationality, quarter and outcome
    # and summing across quarters cannot double count. Verified against the
    # raw file before this script was written.
    counts = defaultdict(lambda: defaultdict(int))
    quarters = defaultdict(set)

    for row in sheet.iter_rows(min_row=5, values_only=True):
        if row[0] is None:
            continue
        if str(row[0]) != UK_TARGET_YEAR or str(row[4]) != UK_VISA_GROUP:
            continue
        nationality = str(row[2])
        if nationality not in wanted:
            continue
        counts[nationality][str(row[8])] += int(row[9] or 0)
        quarters[nationality].add(str(row[1]))

    workbook.close()
    assert_full_coverage({wanted[k]["name"] for k in counts}, "UK Vis_D02")

    records = []
    for nationality in NATIONALITIES:
        outcomes = counts[nationality["uk_label"]]
        issued = outcomes.get("Issued", 0)
        refused = outcomes.get("Refused", 0)
        denominator = issued + refused
        if denominator == 0:
            raise SystemExit(
                f"UK Vis_D02: {nationality['name']} has zero decisions in "
                f"{UK_TARGET_YEAR}. Refusing to write a record with no denominator."
            )
        records.append(
            {
                "nationality": nationality["name"],
                "iso3": nationality["iso3"],
                "destination": source["destination"],
                "visa_group": UK_VISA_GROUP,
                "axis": source["axis"],
                "measure": "refused_share_of_decisions",
                "numerator": refused,
                "denominator": denominator,
                "refusal_rate": round(refused / denominator, 6),
                "refusal_rate_percent": round(refused / denominator * 100, 2),
                "outcome_counts": {
                    "issued": issued,
                    "refused": refused,
                    "withdrawn": outcomes.get("Withdrawn", 0),
                    "lapsed": outcomes.get("Lapsed", 0),
                },
                "quarters_included": sorted(quarters[nationality["uk_label"]]),
                "source_year": source["source_year"],
                "year_basis": source["year_basis"],
                "source_url": source["download_url"],
                "source_landing_url": source["landing_url"],
                "source_table": "Vis_D02",
                "retrieved_at": source["retrieved_at"],
                "methodology": source["methodology"],
                "base_rate_caveat": BASE_RATE_CAVEAT,
            }
        )
    return records


# Matches a table line such as "BAHAMAS, THE 14.46%". The nationality is
# everything before the trailing percentage. Page headers and footers carry no
# percentage and therefore do not match.
US_ROW = re.compile(r"^(?P<name>.+?)\s+(?P<rate>\d{1,3}(?:\.\d{1,2})?)%$")


def parse_us_pdf(path):
    """Return ({nationality label: (rate_percent, published_string, page)}, excluded).

    The table contains one row that is not a nationality:
    "*NON-NATIONALITY BASED ISSUANCES". The footnote on the last page defines
    it as travel documents issued by an authority other than the holder's
    country of nationality, including stateless persons and persons whose
    nationality cannot be determined. It is excluded here so the record count
    is an honest count of nationalities.
    """
    reader = PdfReader(path)
    parsed = {}
    excluded = []
    for page_number, page in enumerate(reader.pages, start=1):
        for line in page.extract_text().splitlines():
            line = line.strip()
            if not line or "ADJUSTED REFUSAL RATE" in line:
                continue
            match = US_ROW.match(line)
            if not match:
                continue
            name = match.group("name").strip()
            if name.startswith("*"):
                excluded.append(name)
                continue
            if name in parsed:
                raise SystemExit(
                    f"US refusal rates: {name} appears more than once in the PDF. "
                    f"Parse is ambiguous, refusing to continue."
                )
            parsed[name] = (
                float(match.group("rate")),
                f"{match.group('rate')}%",
                page_number,
            )
    return parsed, excluded


def build_us(source):
    """US Department of State adjusted refusal rates, B visas, FY2025."""
    parsed, excluded = parse_us_pdf(raw_path(source))
    by_label = {n["us_label"]: n for n in NATIONALITIES}
    found = {by_label[label]["name"] for label in parsed if label in by_label}
    assert_full_coverage(found, "US adjusted refusal rates")

    records = []
    for nationality in NATIONALITIES:
        rate_percent, published, page_number = parsed[nationality["us_label"]]
        records.append(
            {
                "nationality": nationality["name"],
                "iso3": nationality["iso3"],
                "destination": source["destination"],
                "visa_group": "B visitor visas",
                "axis": source["axis"],
                "measure": "adjusted_refusal_rate",
                # The Department publishes the rate only. Reconstructing a
                # numerator and denominator from it would be invention, so
                # both stay null and the methodology explains the omission.
                "numerator": None,
                "denominator": None,
                "refusal_rate": round(rate_percent / 100, 6),
                "refusal_rate_percent": rate_percent,
                "rate_as_published": published,
                "source_page": page_number,
                "source_year": source["source_year"],
                "year_basis": source["year_basis"],
                "source_url": source["download_url"],
                "source_landing_url": source["landing_url"],
                "retrieved_at": source["retrieved_at"],
                "methodology": source["methodology"],
                "base_rate_caveat": BASE_RATE_CAVEAT,
            }
        )
    return records, len(parsed), excluded


# ---------------------------------------------------------------------
# Schengen
# ---------------------------------------------------------------------

# Column indices on both the main consulate sheet and the Cyprus sheet. The
# two sheets label their headers differently, Schengen State versus Member
# State and Total LTVs issued versus Total LTV visas issued, but the positions
# line up, so one set of indices reads both.
SCH_STATE = 0
SCH_COUNTRY = 1
SCH_CONSULATE = 2
SCH_APPLIED = 8
SCH_ISSUED = 9
SCH_MEVS = 10
SCH_LTVS = 12
SCH_NOT_ISSUED = 13

# Trailing rows at the foot of the consulate sheet: Selection Sub total in
# 2025, Total worldwide 2025, and Share of subtotal on worldwide total.
#
# In the 2025 file all three carry an empty Schengen State and country, so the
# emptiness check in _is_data_row already excludes them and this marker list
# never fires. It is kept as redundant defence in case a later edition labels
# a subtotal row while still filling in the country. The real guarantee that
# nothing was miscounted is _cross_check_totals, which reconciles our sums
# against the Commission's own published country totals.
SCH_FOOTER_MARKERS = ("sub total", "subtotal", "total worldwide", "share of subtotal")

SCHENGEN_UNIFORM_VISA = "Schengen uniform short stay visa, category C"
CYPRUS_NATIONAL_VISA = "Cyprus national visa, not a Schengen visa"


def _int(value):
    """Cells arrive as int, float, str or None depending on the sheet."""
    if value is None or value == "":
        return 0
    if isinstance(value, str):
        value = value.strip().replace(",", "")
        if not value:
            return 0
    return int(float(value))


def _clean(value):
    """Collapse whitespace. Names are already correct Unicode.

    An earlier probe reported mojibake in this file, TURKIYE and Santa Fe
    showing replacement characters. That was a Windows console encoding
    artifact, not a data problem: the stored code points are U+00DC and
    U+00E9 and they are correct. Nothing is transliterated here, because
    rewriting a country's name to fix a display bug that does not exist
    would be the actual corruption.
    """
    return " ".join(str(value).split())


def _is_data_row(row):
    if len(row) <= SCH_NOT_ISSUED:
        return False
    if not row[SCH_STATE] or not row[SCH_COUNTRY] or not row[SCH_CONSULATE]:
        return False
    label = str(row[SCH_CONSULATE]).strip().lower()
    return not any(marker in label for marker in SCH_FOOTER_MARKERS)


def _schengen_record(source, level, country, city, state, counts, visa_type, methodology):
    applied = counts["applied"]
    not_issued = counts["not_issued"]
    return {
        "level": level,
        "location_country": country["name"],
        "iso3": country["iso3"],
        "consulate_city": city,
        "schengen_state": state,
        "destination": source["destination"],
        "visa_type": visa_type,
        "axis": source["axis"],
        "measure": "not_issued_share_of_applications",
        "numerator": not_issued,
        "denominator": applied,
        "not_issued_rate": round(not_issued / applied, 6) if applied else None,
        "not_issued_rate_percent": round(not_issued / applied * 100, 2) if applied else None,
        "counts": counts,
        "source_year": source["source_year"],
        "year_basis": source["year_basis"],
        "source_url": source["download_url"],
        "source_landing_url": source["landing_url"],
        "retrieved_at": source["retrieved_at"],
        "methodology": methodology,
        "base_rate_caveat": BASE_RATE_CAVEAT,
    }


def _sum_counts(rows):
    return {
        "applied": sum(_int(r[SCH_APPLIED]) for r in rows),
        "issued": sum(_int(r[SCH_ISSUED]) for r in rows),
        "mevs_issued": sum(_int(r[SCH_MEVS]) for r in rows),
        "ltvs_issued": sum(_int(r[SCH_LTVS]) for r in rows),
        "not_issued": sum(_int(r[SCH_NOT_ISSUED]) for r in rows),
    }


def _cross_check_totals(workbook, by_country):
    """Validate our row filtering against the source's own published totals.

    The Totals by third country sheet carries the Commission's own figure for
    each country. If our consulate rows sum to something different, we either
    dropped a real row or kept a footer row, and the build should stop.
    """
    # This sheet carries a blank row and a header before the data, and the
    # offset has moved between editions, so rows are filtered on being
    # numeric rather than on a hardcoded starting row.
    sheet = workbook["Totals - third country"]
    published = {}
    for row in sheet.iter_rows(values_only=True):
        if len(row) < 8 or not row[1]:
            continue
        try:
            published[_clean(row[1]).upper()] = (_int(row[2]), _int(row[6]))
        except (TypeError, ValueError):
            continue

    for country, rows in by_country:
        label = country["schengen_label"]
        if label not in published:
            continue
        counts = _sum_counts(rows)
        applied, not_issued = published[label]
        if (counts["applied"], counts["not_issued"]) != (applied, not_issued):
            raise SystemExit(
                f"Schengen: {country['name']} does not reconcile against the "
                f"published country total. Parsed applied={counts['applied']} "
                f"not_issued={counts['not_issued']}, published applied={applied} "
                f"not_issued={not_issued}. Check the footer row filter."
            )


def build_schengen(source):
    """EU Commission Schengen consulate statistics, 2025.

    Emits three levels for each of the 12 countries. A single consulate, for
    example France at Lagos, sits alongside the aggregate for all Schengen
    states at Lagos, and alongside the aggregate for the whole country.
    """
    wanted = {c["schengen_label"]: c for c in COUNTRIES}
    workbook = openpyxl.load_workbook(raw_path(source), read_only=True, data_only=True)

    sheet = workbook["Data for consulates"]
    by_country = defaultdict(list)
    total_rows = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not _is_data_row(row):
            continue
        total_rows += 1
        label = _clean(row[SCH_COUNTRY]).upper()
        if label in wanted:
            by_country[wanted[label]["name"]].append(row)

    found = {name for name in by_country}
    assert_full_coverage(found, "Schengen consulates")

    _cross_check_totals(workbook, [(c, by_country[c["name"]]) for c in COUNTRIES])

    records = []
    for country in COUNTRIES:
        rows = by_country[country["name"]]

        # Country wide, every Schengen state at every city in that country.
        records.append(
            _schengen_record(
                source, "consulate_country", country, None, None,
                _sum_counts(rows), SCHENGEN_UNIFORM_VISA, source["methodology"],
            )
        )

        by_city = defaultdict(list)
        for row in rows:
            by_city[_clean(row[SCH_CONSULATE])].append(row)

        for city in sorted(by_city):
            # Every Schengen state issuing in that city.
            records.append(
                _schengen_record(
                    source, "consulate_city", country, city, None,
                    _sum_counts(by_city[city]), SCHENGEN_UNIFORM_VISA, source["methodology"],
                )
            )
            # One record per issuing state, so France at Lagos stands alone.
            for row in sorted(by_city[city], key=lambda r: _clean(r[SCH_STATE])):
                records.append(
                    _schengen_record(
                        source, "consulate", country, city, _clean(row[SCH_STATE]),
                        _sum_counts([row]), SCHENGEN_UNIFORM_VISA, source["methodology"],
                    )
                )

    # Cyprus sits on its own sheet because it did not fully apply the Schengen
    # acquis in 2025, and the source says its figures are national visas. Those
    # do not describe Schengen access, so they are emitted as their own records
    # and deliberately kept out of every aggregate above.
    cyprus_records = []
    for row in workbook["Cyprus"].iter_rows(min_row=2, values_only=True):
        if not _is_data_row(row):
            continue
        label = _clean(row[SCH_COUNTRY]).upper()
        if label not in wanted:
            continue
        cyprus_records.append(
            _schengen_record(
                source, "consulate", wanted[label], _clean(row[SCH_CONSULATE]),
                _clean(row[SCH_STATE]), _sum_counts([row]),
                CYPRUS_NATIONAL_VISA, SCHENGEN_CYPRUS_METHODOLOGY,
            )
        )

    workbook.close()
    return records, cyprus_records, total_rows


def build_financial(source):
    """Validate the curated financial requirement table and attach provenance.

    Nothing is parsed here. The records are transcribed by hand in
    scripts/financial_requirements.py because the source is prose, so this
    function's whole job is to refuse to write anything that is internally
    inconsistent or incomplete.
    """
    methodologies = {
        "Schengen area": SCHENGEN_FINANCIAL_METHODOLOGY,
        "United Kingdom": UK_FINANCIAL_METHODOLOGY,
        "United States": US_FINANCIAL_METHODOLOGY,
    }

    records = []
    for record in FINANCIAL_RECORDS:
        record = dict(record)
        state = record["state"]

        if record["basis"] not in BASES:
            raise SystemExit(f"financial: {state} has unknown basis {record['basis']!r}.")
        if record["amount_status"] not in AMOUNT_STATUSES:
            raise SystemExit(
                f"financial: {state} has unknown amount_status {record['amount_status']!r}."
            )
        if record["jurisdiction"] not in methodologies:
            raise SystemExit(f"financial: {state} has unknown jurisdiction.")
        if not record["source_url"] or not record["source_year"]:
            raise SystemExit(f"financial: {state} is missing source_url or source_year.")

        amounts = [
            record["per_day_amount"],
            record["per_entry_amount"],
            record["trip_minimum_amount"],
        ]
        published = record["basis"] != "none_published"

        # An amount and a claim that nothing is published cannot both be true.
        if not published:
            if any(a is not None for a in amounts) or record["currency"] is not None:
                raise SystemExit(
                    f"financial: {state} says none_published but carries an amount. "
                    f"One of the two is wrong."
                )
            if record["amount_status"] != "none_published":
                raise SystemExit(
                    f"financial: {state} says none_published but amount_status is "
                    f"{record['amount_status']!r}."
                )
        else:
            if record["currency"] is None:
                raise SystemExit(f"financial: {state} publishes an amount but no currency.")
            if all(a is None for a in amounts) and not record["variants"]:
                raise SystemExit(
                    f"financial: {state} claims a published requirement but records "
                    f"no amount anywhere, not even in variants."
                )
        for value in amounts:
            if value is not None and value <= 0:
                raise SystemExit(f"financial: {state} has a non positive amount.")

        record["axis"] = "destination"
        record["measure"] = "published_financial_requirement"
        record["retrieved_at"] = FINANCIAL_RETRIEVED_AT
        record["methodology"] = methodologies[record["jurisdiction"]]
        record["threshold_caveat"] = FINANCIAL_THRESHOLD_CAVEAT
        records.append(record)

    # Coverage. Every Schengen state must be present exactly once.
    found = {r["state"] for r in records if r["in_schengen_area"]}
    missing = [s for s in SCHENGEN_STATES if s not in found]
    if missing:
        raise SystemExit(
            f"financial: no record for {', '.join(missing)}. Every Schengen state "
            f"needs a record, including one that says no amount is published. "
            f"Add it to scripts/financial_requirements.py."
        )
    unexpected = sorted(found - set(SCHENGEN_STATES))
    if unexpected:
        raise SystemExit(
            f"financial: {', '.join(unexpected)} is flagged as in the Schengen "
            f"area but is not in SCHENGEN_STATES."
        )
    states = [r["state"] for r in records]
    duplicates = sorted({s for s in states if states.count(s) > 1})
    if duplicates:
        raise SystemExit(f"financial: duplicate records for {', '.join(duplicates)}.")

    for jurisdiction in ("United Kingdom", "United States"):
        if not any(r["jurisdiction"] == jurisdiction for r in records):
            raise SystemExit(
                f"financial: no record for {jurisdiction}. The absence of a "
                f"published threshold is a record, not a reason to omit one."
            )
    return records


def write_financial(records):
    published = [r for r in records if r["basis"] != "none_published"]
    silent = [r for r in records if r["basis"] == "none_published"]
    schengen_records = [r for r in records if r["jurisdiction"] == "Schengen area"]
    checked = [r for r in schengen_records if r.get("national_source")]
    disagree = [
        r["state"] for r in checked if r["national_source"]["agrees_with_annex"] is False
    ]

    payload = {
        "schema_version": 1,
        "dataset_id": "published_financial_requirements",
        "title": "Published financial requirements for the three supported destinations",
        "description": (
            "What each destination publishes as the money a short stay visitor "
            "must show. Schengen states publish a per day reference amount "
            "each. The United Kingdom and the United States publish no figure "
            "and assess adequacy case by case."
        ),
        "destinations": ["Schengen area", "United Kingdom", "United States"],
        "axis": "destination",
        "axis_warning": (
            "This dataset is keyed by the destination that sets the "
            "requirement. It is not keyed by passport and not keyed by where "
            "the application is made. Do not join it to the refusal datasets "
            "as though the three shared an axis, and do not present a "
            "financial requirement as though it moved the refusal rate."
        ),
        "source_year": 2026,
        "year_basis": "publication_version_date",
        "retrieved_at": FINANCIAL_RETRIEVED_AT,
        "annex_version": ANNEX_VERSION,
        "methodology_by_jurisdiction": {
            "Schengen area": SCHENGEN_FINANCIAL_METHODOLOGY,
            "United Kingdom": UK_FINANCIAL_METHODOLOGY,
            "United States": US_FINANCIAL_METHODOLOGY,
        },
        "threshold_caveat": FINANCIAL_THRESHOLD_CAVEAT,
        "curation_note": (
            "Transcribed by hand from prose, not parsed. The Commission annex "
            "gives one written section per state rather than a table, so the "
            "reading is recorded in scripts/financial_requirements.py where it "
            "can be checked against the source line by line."
        ),
        "generated_by": "scripts/build-dataset.py",
        "record_count": len(records),
        "records_with_published_amount": len(published),
        "records_with_no_published_amount": len(silent),
        "schengen_states_total": len(schengen_records),
        "schengen_states_checked_against_national_page": len(checked),
        "schengen_states_where_national_page_disagrees": disagree,
        "not_found": [],
        "not_found_note": (
            "Empty, and that is a real result rather than an unchecked box. "
            "Every Schengen state has a section in the Commission annex, which "
            "is an official EU publication of the amounts the states "
            "themselves notified, so no state had to be recorded as not found. "
            "Read schengen_states_checked_against_national_page next to it: "
            "most figures rest on that one compilation, and only the states "
            "carrying a national_source were also read on their own "
            "government's page. Luxembourg is the case worth knowing about, "
            "since its own page publishes no amount at all."
        ),
        "records": records,
    }

    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    path = PROCESSED_DIR / "financial-requirements.json"
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    return path


def write_dataset(filename, source, records, extra_header=None):
    payload = {
        "schema_version": 1,
        "dataset_id": source["id"],
        "title": source["title"],
        "description": source["description"],
        "destination": source["destination"],
        "axis": source["axis"],
        "source_year": source["source_year"],
        "year_basis": source["year_basis"],
        "source_url": source["download_url"],
        "source_landing_url": source["landing_url"],
        "retrieved_at": source["retrieved_at"],
        "publication": source["publication"],
        "methodology": source["methodology"],
        "base_rate_caveat": BASE_RATE_CAVEAT,
        "generated_by": "scripts/build-dataset.py",
        "record_count": len(records),
    }
    if extra_header:
        payload.update(extra_header)
    payload["records"] = records

    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    path = PROCESSED_DIR / filename
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    return path


def report(title, records, show_counts):
    print(f"\n{title}")
    if show_counts:
        print(f"  {'nationality':13} {'refused':>8} {'decisions':>10} {'rate':>8}")
        for record in sorted(records, key=lambda r: -r["refusal_rate"]):
            print(
                f"  {record['nationality']:13} {record['numerator']:>8} "
                f"{record['denominator']:>10} {record['refusal_rate_percent']:>7.2f}%"
            )
    else:
        print(f"  {'nationality':13} {'rate':>8}  {'published':>10}  page")
        for record in sorted(records, key=lambda r: -r["refusal_rate"]):
            print(
                f"  {record['nationality']:13} {record['refusal_rate_percent']:>7.2f}%  "
                f"{record['rate_as_published']:>10}  p{record['source_page']}"
            )


def main():
    manifest = load_manifest()

    uk_source = manifest[UK_SOURCE_ID]
    uk_records = build_uk(uk_source)
    uk_path = write_dataset(
        "uk-visitor-refusals.json",
        uk_source,
        uk_records,
        extra_header={"visa_group": UK_VISA_GROUP, "source_table": "Vis_D02"},
    )

    us_source = manifest[US_SOURCE_ID]
    us_records, us_total_rows, us_excluded = build_us(us_source)
    us_path = write_dataset(
        "us-b-visa-refusals.json",
        us_source,
        us_records,
        extra_header={
            "visa_group": "B visitor visas",
            "nationalities_in_source": us_total_rows,
            "non_nationality_rows_excluded": us_excluded,
        },
    )

    report(
        f"UK, Visitor group, calendar year {UK_TARGET_YEAR}, per decision",
        uk_records,
        show_counts=True,
    )
    report(
        "US, B visas, fiscal year 2025, per person adjusted rate",
        us_records,
        show_counts=False,
    )

    schengen_source = manifest[SCHENGEN_SOURCE_ID]
    sch_records, cyprus_records, sch_total_rows = build_schengen(schengen_source)
    all_schengen = sch_records + cyprus_records
    schengen_path = write_dataset(
        "schengen-consulate-not-issued.json",
        schengen_source,
        all_schengen,
        extra_header={
            "visa_type": SCHENGEN_UNIFORM_VISA,
            "levels": ["consulate_country", "consulate_city", "consulate"],
            "consulate_rows_in_source": sch_total_rows,
            "cyprus_records": len(cyprus_records),
            "cyprus_note": SCHENGEN_CYPRUS_METHODOLOGY,
            "axis_warning": (
                "This dataset is keyed by where an application was made, not "
                "by the applicant's passport. Never render it as a refusal "
                "rate for a nationality, and never place it on a shared axis "
                "with the UK or US datasets without saying what each measures."
            ),
        },
    )

    print("\nSchengen, calendar year 2025, uniform visas not issued, by consulate location")
    print(f"  {'country':13} {'not issued':>11} {'applied':>10} {'rate':>8}  cities")
    for record in sorted(
        (r for r in sch_records if r["level"] == "consulate_country"),
        key=lambda r: -r["not_issued_rate"],
    ):
        cities = sorted(
            {
                r["consulate_city"]
                for r in sch_records
                if r["level"] == "consulate_city" and r["location_country"] == record["location_country"]
            }
        )
        print(
            f"  {record['location_country']:13} {record['numerator']:>11} "
            f"{record['denominator']:>10} {record['not_issued_rate_percent']:>7.2f}%  {', '.join(cities)}"
        )

    financial_records = build_financial(manifest[FINANCIAL_SOURCE_ID])
    financial_path = write_financial(financial_records)

    print("\nPublished financial requirements, by destination")
    print(f"  {'state':16} {'basis':28} {'per day':>10}  {'status':24} national page")
    for record in financial_records:
        if record["per_day_amount"] is None:
            amount = "none" if record["basis"] == "none_published" else "see variants"
        else:
            amount = f"{record['per_day_amount']:.2f} {record['currency']}"
        national = "not checked"
        if record.get("national_source"):
            source = record["national_source"]
            if not source["states_an_amount"]:
                national = "checked, publishes no amount"
            elif source["agrees_with_annex"]:
                national = "checked, agrees"
            else:
                national = "checked, DISAGREES"
        print(
            f"  {record['state']:16} {record['basis']:28} {amount:>10}  "
            f"{record['amount_status']:24} {national}"
        )

    print(f"\nwrote {uk_path.relative_to(REPO_ROOT)}  ({len(uk_records)} records)")
    print(f"wrote {us_path.relative_to(REPO_ROOT)}  ({len(us_records)} records)")
    print(f"wrote {schengen_path.relative_to(REPO_ROOT)}  ({len(all_schengen)} records)")
    print(f"US PDF contained {us_total_rows} nationalities, 12 selected.")
    if us_excluded:
        print(f"US PDF non nationality rows excluded: {', '.join(us_excluded)}")
    print(f"Schengen sheet had {sch_total_rows} consulate rows after dropping footers.")
    print(f"Schengen records: {len(sch_records)} uniform visa, {len(cyprus_records)} Cyprus national visa.")
    print(f"wrote {financial_path.relative_to(REPO_ROOT)}  ({len(financial_records)} records)")
    silent = [r["state"] for r in financial_records if r["basis"] == "none_published"]
    print(f"Destinations publishing no financial threshold: {', '.join(silent)}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
